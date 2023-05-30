from openpyxl.reader.excel import load_workbook

from formatter.table_formatter import BaseTableFormatter
from openpyxl.styles import Alignment, PatternFill
import json


class FsXlsTableFormatter(BaseTableFormatter):

    @classmethod
    def format(cls, table_path, format_delimiter):
        workbook = load_workbook(table_path)
        worksheet = workbook['Sheet1']
        for row in worksheet:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
                if isinstance(cell.value, str):
                    values = cell.value.split(sep=format_delimiter)
                    if len(values) == 2:
                        features_result = json.loads(values[1])
                        cell.fill = cls.calculate_color(features_result.get('sa', 0))

        columns = [col for col in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'] + ['AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH']
        for column in columns:
            worksheet.column_dimensions[column].width = cls.width

        workbook.save(table_path)

    @staticmethod
    def calculate_color(sentiment: float) -> PatternFill:
        rgb_color = [255, 255, 255]

        if sentiment != 0:
            if sentiment > 0:
                rgb_color[0] = rgb_color[2] = int(255 * (1 - sentiment))
            else:
                rgb_color[1] = rgb_color[2] = int(255 * (1 + sentiment))

        # Convert RGB to hexadecimal format
        hex_color = '%02x%02x%02x' % tuple(rgb_color)
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        return fill
